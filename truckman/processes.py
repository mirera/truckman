from django.utils import timezone
from datetime import timedelta
from trip.models import Estimate, Invoice, LoadingList, LoadingListItem, DailyRegister, Load, DailyRegister, Driver, Trip, EstimateItem, InvoiceItem
from django.shortcuts import get_object_or_404
from truckman.utils import send_email, send_whatsapp_text, get_location_data
from django.db.models import F
import pandas as pd
from openpyxl import Workbook
import pytz
from collections import defaultdict
from authentication.models import WhatsappSetting

'''
Function to get a trip vehicle(s)
'''
def get_trip_vehicles(trip):
    loads = Load.objects.filter(estimate=trip.estimate)
    # Get all vehicles assigned to loads
    vehicles = []
    for load in loads:
        if load.assigned_truck:
            vehicles.append(load.assigned_truck) 
    return vehicles
#--ends

'''
#helper function to calculate due date 
of an invoice based on the customer
'''
def get_invoice_due_date(customer, trip):
    payment_terms = {
        '2 Days': 2,
        '7 Days': 7,
        '10 Days': 10,
        '15 Days': 15,
        '30 Days': 30,
        'Cash on Delivery': None,  # No need to calculate a due date for Cash on Delivery
    }

    payment_term = customer.payment_term
    if payment_term in payment_terms:
        days = payment_terms[payment_term]
        if days is not None:
            due_date = trip.start_time.date() + timedelta(days=days)
        else:
            due_date = trip.end_time.date()  # For 'Cash on Delivery'
    else:
        # Handle cases where payment term is not recognized
        due_date = None  

    return due_date

'''
This function is called when a estimate has been acceptes
by client. You can also use it to re-generate an 
invoice for a given estimate.
'''
def generate_invoice(estimate): 
        trip = Trip.objects.get(estimate=estimate)
        due_date = get_invoice_due_date(estimate.customer, trip)

        #get estimate items
        estimate_items = EstimateItem.objects.filter(estimate=estimate)

        #create invoice instance from estimate data
        invoice = Invoice.objects.create(
            company=estimate.company,
            estimate = estimate,
            sub_total = estimate.sub_total,
            discount = estimate.discount,
            tax = estimate.tax,
            total = estimate.total,
            balance = estimate.total,
            invoice_date = timezone.now().date(),
            due_date = due_date,
            note = estimate.company.invoice_payment_details,
        )

        #create invoice items from estimate items
        invoice_items = []

        for item in estimate_items:
            invoice_item = InvoiceItem.objects.create(
                    company=estimate.company,
                    invoice=invoice,
                    route = item.route,
                    item_type = item.item_type,
                    trucks = item.trucks,
                    rate = item.rate,
                    amount = item.amount,
                    description = item.description,
                )
            invoice_items.append(invoice_item)
    
        return invoice       
#--ends

'''
This function is called when a loas has been assigned a truck
It creates a loading list 
'''
def generate_loading_list(estimate):
        # Get all vehicles assigned to loads
        trip = Trip.objects.get(estimate=estimate)
        vehicles = get_trip_vehicles(trip)
        
        #create instance of a loading list
        loading_list = LoadingList.objects.create(
            company=estimate.company,
            estimate = estimate,    
        )

        # Create an instance of a loading list item for each vehicle
        loading_list_items = []
        for vehicle in vehicles:
            driver = get_object_or_404(Driver, assigned_vehicle=vehicle)
            
            loading_list_item = LoadingListItem.objects.create(
                company=estimate.company,
                loading_list=loading_list,
                vehicle=vehicle,
                driver=driver,
            )
            loading_list_items.append(loading_list_item)

        return loading_list
#--ends     

'''
funtion to create a daily register instant
'''
def create_daily_register(request, company, trip, vehicle, driver, coordinates, user_timezone, datetime_field):
    # Get user sublocality, city, and country
    nearest_town, country = get_location_data(coordinates)

    # Create a DailyRegister object
    return DailyRegister.objects.create(
        company=company,
        trip=trip,
        vehicle=vehicle,
        driver=driver,
        **{
            f'{datetime_field}_location': nearest_town,
            'country': country,
            f'{datetime_field}_odemeter_reading': request.POST.get(f'{datetime_field}_odemeter_reading'),
            'vehicle_status': request.POST.get('vehicle_status'),
            'reason_parked': request.POST.get('reason_parked'),
        }
    )

#--end

'''
Function to get register entries of a day
i.e morning, midday and evening
'''
def get_driver_day_entries(trip, driver, target_date):
    day_entries = DailyRegister.objects.filter(
        driver=driver,
        trip=trip,
        submission_time__date=target_date
    )
    return day_entries
#--ends

'''
Function to get register daily entries
all entries of a load/vehicle on a trip
'''
def get_load_daily_entries(load):
    trip = Trip.objects.get(estimate=load.estimate) 
    load_entries = DailyRegister.objects.filter(
        trip=trip,
        vehicle=load.assigned_truck
    )
    return load_entries
#--ends

'''
Split the entries into morning, midday, and evening based on submission times
'''
def split_day_entry(entry, request):
    user_timezone = request.session.get('user_timezone', 'UTC') #get user timezone 
    user_tz = pytz.timezone(user_timezone)# Create a timezone object
    submission_time = entry.submission_time.astimezone(user_tz) # Convert UTC time to the user's local time
    morning_entry, midday_entry, evening_entry = None, None, None
    morning_sub_time, mid_sub_time, evening_sub_time = None, None, None
    
    if 6 <= submission_time.hour < 12:
        # Convert sub_time UTC time to the user's local time format to HMS
        morning_sub_time = entry.submission_time.astimezone(user_tz).strftime('%H:%M:%S')
        morning_entry = entry
    elif 12 <= submission_time.hour < 16:
        mid_sub_time = entry.submission_time.astimezone(user_tz).strftime('%H:%M:%S')
        midday_entry = entry
    elif 16 <= submission_time.hour < 21:
        evening_sub_time = entry.submission_time.astimezone(user_tz).strftime('%H:%M:%S')
        evening_entry = entry

    return morning_entry, midday_entry, evening_entry, morning_sub_time, mid_sub_time, evening_sub_time
#--ends


def create_sheetdata(request, worksheet, vehicle, entries_for_date, n, m, p, q, x):
    morning = None
    midday = None
    evening = None
    morning_time = None
    midday_time = None
    evening_time = None

    for entry in entries_for_date:
        #split a day's entry into morning, mid and evening
        morning_entry, midday_entry, evening_entry, morning_sub_time, mid_sub_time, evening_sub_time = split_day_entry(entry, request)

        if morning_entry:
            morning = morning_entry
            morning_time = morning_sub_time
        if midday_entry:
            midday = midday_entry
            midday_time = mid_sub_time
        if evening_entry:
            evening = evening_entry
            evening_time = evening_sub_time


    #Create rows/columns and insert sheet data
    worksheet['A1'] = f'{vehicle.plate_number} Report'
    if entries_for_date:
        worksheet[f'A{n}'] = f'Date | {entries_for_date[0].submission_time.date()}' 
    else:
        worksheet[f'A{n}'] = f'Date | Not Data Available'

    worksheet[f'B{n}'] = "#"
    worksheet[f'C{n}'] = "Morning"
    worksheet[f'D{n}'] = "Midday"
    worksheet[f'E{n}'] = "Evening"
    worksheet[f'B{m}'] = "Submission Time"
    worksheet[f'B{p}'] = "Location"
    worksheet[f'B{q}'] = "Odemeter"
    worksheet[f'B{x}'] = "Truck Status"
    
    if not morning:
        worksheet[f'C{m}'] = 'N/A'
        worksheet[f'C{p}'] = 'N/A'
        worksheet[f'C{q}'] = 'N/A'
        worksheet[f'C{x}'] = 'N/A'
    else:
        worksheet[f'C{m}'] = morning_time #morning entry subtime
        worksheet[f'C{p}'] = morning.morning_location #morning entry location
        worksheet[f'C{q}'] = morning.morning_odemeter_reading#morning odemeter reading
        worksheet[f'C{x}'] = morning.vehicle_status #morning trip load status
    
    if not midday:
        worksheet[f'D{m}'] = 'N/A'
        worksheet[f'D{p}'] = 'N/A'
        worksheet[f'D{q}'] = 'N/A'
        worksheet[f'D{x}'] = 'N/A'
    else:
        worksheet[f'D{m}'] = midday_time#midday entry subtime
        worksheet[f'D{p}'] = midday.midday_location #midday entry location
        worksheet[f'D{q}'] = midday.midday_odemeter_reading #midday odemeter reading
        worksheet[f'D{x}'] = midday.vehicle_status #midday trip load status
    
    if not evening:
        worksheet[f'E{m}'] = 'N/A'
        worksheet[f'E{p}'] = 'N/A'
        worksheet[f'E{q}'] = 'N/A'
        worksheet[f'E{x}'] = 'N/A'
    else:
        worksheet[f'E{m}'] = evening_time #evening entry subtime
        worksheet[f'E{p}'] = evening.evening_location #evening entry location
        worksheet[f'E{q}'] = evening.evening_odemeter_reading #evening odemeter reading
        worksheet[f'E{x}'] = evening.vehicle_status #evening trip load status
 
#--ends

def create_workbook(request, trip):
    '''
    Function to create a workbook
    '''
    workbook = Workbook()
    #create a worksheets for each vehicle on the trip
    for vehicle in get_trip_vehicles(trip):
        worksheet = workbook.create_sheet(title=f"{trip.trip_id} {vehicle.plate_number} Report")
        load = get_object_or_404(Load, assigned_truck=vehicle, status='On Transit') #find a way to accomodate  completed and on transit loads
        entries = get_load_daily_entries(load)#all entries made for this load

        #group the entries into a list of lists of entries with the same date.
        grouped_entries = defaultdict(list)
        for entry in entries:
            grouped_entries[entry.submission_time.date()].append(entry)
        
        n,m,p,q,x = 2,3,4,5,6 # these variables dynamically help write on excel vertically
        for date, entries_for_date in grouped_entries.items():#create sheet data for each day
                create_sheetdata(request, worksheet, vehicle, entries_for_date, n, m, p, q, x) 
                n += 6
                m += 6
                p += 6
                q += 6
                x += 6
    # Save the file
    #workbook.save(f'daily_reports/{trip.trip_id} Daily Report.xlsx') 
    return workbook
#--ends 

 
'''
Generate single day daily report for client sharing
'''
def generate_client_daily_report(trip, path_to_workbook):
    trip_workbook = create_workbook(path_to_workbook)#create an Excel workbook
    worksheet = create_worksheet(path_to_workbook, trip)#create worksheet
    create_sheetdata(path_to_workbook, trip) #write or load data into to the worksheet
    return trip_workbook #this should be a path to the workbook


#send trip ,dispatched, vehicles tracking uri to client
def send_trip_vehicle_tURI(trip):
    company = trip.company
    if trip.status == 'Dispatched':
        customer = trip.estimate.customer #get customer 
        vehicles = get_trip_vehicles(trip) #get trip vehicles
        for vehicle in vehicles:
            whatsapp_setting = get_object_or_404(WhatsappSetting, company=company ) 
            load = get_object_or_404(Load, assigned_truck=vehicle)
            tracking_uri = vehicle.tracking_uri #later on make this url expire when the load status is delivered. 
            message = f'Dear {company.name}, your load {load.load_id} on truck of plate number {vehicle.plate_number} has been loaded and dispatched. Here is the tracking link {tracking_uri}.'
            send_whatsapp_text(
                instance_id=whatsapp_setting.instance_id, 
                access_token=whatsapp_setting.access_token, 
                phone_no=customer.phone, 
                message=message
            )
#--ends


'''
This updates trip status to dispatched or completed dep
depening on load status
'''
def update_trip_status(load):
    trip = Trip.objects.get(estimate=load.estimate)
    
    if trip:
        # Retrieve all loads associated with the trip
        associated_loads = Load.objects.filter(estimate=load.estimate)
        
        # Check if any load status is 'On Transit'
        if associated_loads.filter(status='On Transit').exists():
            trip.status = 'Dispatched'
            trip.start_time = timezone.now()
            trip.save()
            #send client vehicle tracking uris
            send_trip_vehicle_tURI(trip)
        
        # Check if all loads' status are 'Delivered'
        if associated_loads.exclude(status='Delivered').exists():
            # Not all loads are delivered
            return
        
        # All loads are delivered
        trip.status = 'Completed'
        trip.end_time = timezone.now()
        trip.save()

        # Change all vehicle is_available to True
        vehicles = []
        for load in associated_loads:
            if load.assigned_truck:
                vehicles.append(load.assigned_truck)

        for vehicle in vehicles:
            vehicle.is_available = True
            vehicle.save()
#--ends



